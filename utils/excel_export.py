"""Excel export helper."""
from __future__ import annotations

from datetime import datetime
from typing import List


def generate_excel_file_for_period(news_items: List[dict]) -> str | None:
    """Generate Excel file for news items list."""
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        import tempfile

        wb = Workbook()
        ws = wb.active
        ws.title = "News"

        headers = [
            "Дата",
            "Время",
            "Источник",
            "Ссылка",
            "Заголовок",
            "Содержание новости",
            "Хештеги",
        ]
        ws.append(headers)

        category_map = {
            'world': '#Мир',
            'russia': '#Россия',
            'moscow': '#Москва',
            'moscow_region': '#Подмосковье',
        }

        for news in news_items:
            content = news.get('ai_summary') or news.get('clean_text') or news.get('lead_text') or ""
            content = str(content).strip()
            # Full hierarchical list (g0, g1?, g2?, g3?, r0); prefer stored hashtags
            tags_full = (news.get('hashtags') or news.get('hashtags_ru') or news.get('hashtags_en') or "").strip()
            tag = tags_full or category_map.get(news.get('category', 'russia'), '#Россия')
            published_date = news.get('published_date')
            published_time = news.get('published_time')
            if not published_date and news.get('published_at'):
                try:
                    pub = str(news.get('published_at')).replace('Z', '+00:00')
                    dt = datetime.fromisoformat(pub)
                    published_date = dt.strftime('%Y-%m-%d')
                    published_time = dt.strftime('%H:%M')
                except Exception:
                    published_date = ''
                    published_time = ''
            ws.append([
                published_date or '',
                published_time or '',
                news.get('source', ''),
                news.get('url', ''),
                news.get('title', ''),
                content,
                tag,
            ])

        # Set column widths for readability
        col_widths = [14, 10, 25, 50, 60, 80, 15]
        for i, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width

        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        temp_file.close()
        return temp_file.name
    except Exception:
        return None
