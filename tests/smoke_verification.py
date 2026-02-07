"""Smoke verification for ingestion/date/pause/excel."""
from __future__ import annotations

import os
import sys
import tempfile

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)

from db.database import NewsDatabase
from utils.date_parser import parse_published_at
from utils.content_quality import content_quality_score, is_low_quality


def _assert(condition: bool, message: str) -> None:
    if not condition:
        raise AssertionError(message)


def test_published_at_parsing() -> None:
    html = """
    <html><head>
      <meta property="article:published_time" content="2026-02-07T12:34:56+00:00" />
    </head><body></body></html>
    """
    dt = parse_published_at(html, "https://example.com/2026/02/07/test")
    _assert(dt is not None, "meta published_time should parse")

    html_time = "<time datetime=\"2026-02-07T05:06:07\"></time>"
    dt2 = parse_published_at(html_time, "https://example.com/test")
    _assert(dt2 is not None, "time tag should parse")

    html_jsonld = """
    <script type="application/ld+json">{"@type":"NewsArticle","datePublished":"2026-02-07T01:02:03Z"}</script>
    """
    dt3 = parse_published_at(html_jsonld, "https://example.com/test")
    _assert(dt3 is not None, "JSON-LD datePublished should parse")


def test_quality_gate() -> None:
    score, _meta = content_quality_score("Короткий текст", "Заголовок")
    _assert(is_low_quality(score), "short text should be low quality")


def test_pause_version_gate() -> None:
    temp_db = os.path.join(tempfile.gettempdir(), "news_smoke.db")
    if os.path.exists(temp_db):
        os.remove(temp_db)
    db = NewsDatabase(db_path=temp_db)

    user_id = "1001"
    db.set_pause_state(user_id, False)
    state = db.get_delivery_state(user_id)
    _assert(state["pause_version"] >= 1, "pause_version should be initialized")

    before = state["pause_version"]
    db.set_pause_state(user_id, True)
    after = db.get_delivery_state(user_id)["pause_version"]
    _assert(after == before + 1, "pause_version should increment on pause")

    inserted = db.try_log_delivery(user_id, 1)
    _assert(inserted, "first delivery log insert should succeed")
    inserted2 = db.try_log_delivery(user_id, 1)
    _assert(not inserted2, "duplicate delivery log insert should be blocked")


def test_excel_columns() -> None:
    from openpyxl import load_workbook
    from utils.excel_export import generate_excel_file_for_period

    sample = [
        {
            "id": 1,
            "url": "https://example.com",
            "title": "Test",
            "source": "example.com",
            "category": "world",
            "lead_text": "Text",
            "clean_text": "Text",
            "ai_summary": None,
            "published_date": "2026-02-07",
            "published_time": "12:00",
        }
    ]
    path = generate_excel_file_for_period(sample)
    _assert(path is not None, "excel file should be created")

    wb = load_workbook(path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    expected = ["Дата", "Время", "Источник", "Ссылка", "Заголовок", "Содержание новости", "Хештэг"]
    _assert(headers == expected, "excel headers should match expected columns")
    wb.close()
    os.remove(path)


def main() -> None:
    test_published_at_parsing()
    test_quality_gate()
    test_pause_version_gate()
    test_excel_columns()
    print("SMOKE OK")


if __name__ == "__main__":
    main()
